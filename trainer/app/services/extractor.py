import csv
import logging
import os

logger = logging.getLogger(__name__)


def extract_text(file_path: str) -> str:
    """Extract text content from a file based on its extension.

    Supported formats: PDF, DOCX, XLSX, CSV, TXT, MD.
    Saves extracted text to {file_path}.extracted.txt alongside the original.
    """
    if not os.path.exists(file_path):
        logger.error("File not found: %s", file_path)
        return ""

    ext = os.path.splitext(file_path)[1].lower()
    text = ""

    try:
        if ext == ".pdf":
            text = _extract_pdf(file_path)
        elif ext == ".docx":
            text = _extract_docx(file_path)
        elif ext == ".xlsx":
            text = _extract_xlsx(file_path)
        elif ext == ".csv":
            text = _extract_csv(file_path)
        elif ext in (".txt", ".md"):
            text = _extract_plain(file_path)
        else:
            logger.warning("Unknown extension '%s', attempting plain text read for: %s", ext, file_path)
            text = _extract_plain(file_path)
    except Exception:
        logger.exception("Error extracting text from %s", file_path)
        return ""

    extracted_path = f"{file_path}.extracted.txt"
    try:
        with open(extracted_path, "w", encoding="utf-8") as f:
            f.write(text)
        logger.info("Extracted text saved to %s (%d chars)", extracted_path, len(text))
    except Exception:
        logger.exception("Failed to save extracted text to %s", extracted_path)

    return text


def _extract_pdf(file_path: str) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(file_path)
    pages_text = []
    for i, page in enumerate(reader.pages):
        page_text = page.extract_text()
        if page_text:
            pages_text.append(page_text)
        else:
            logger.debug("PDF page %d had no extractable text", i)
    return "\n\n".join(pages_text)


def _extract_docx(file_path: str) -> str:
    from docx import Document

    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Hoja: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            line = "\t".join(cells).strip()
            if line:
                parts.append(line)
    wb.close()
    return "\n".join(parts)


def _extract_csv(file_path: str) -> str:
    parts = []
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            line = "\t".join(row).strip()
            if line:
                parts.append(line)
    return "\n".join(parts)


def _extract_plain(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()
